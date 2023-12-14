import openpyxl
file="C:\Automation Testing\TestData1.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
colms=sheet.max_column
#for r in range(1,7) :
  #  for c in range(1,4):
   #     sheet.cell(r,c).value ="welcome"
sheet.cell(1,1).value="123"
sheet.cell(1,2).value="Ayaan"
sheet.cell(1,3).value="Doctor"


sheet.cell(2,1).value="126"
sheet.cell(2,2).value="Arun"
sheet.cell(2,3).value="Engineer"


sheet.cell(3,1).value="145"
sheet.cell(3,2).value="Harsh"
sheet.cell(3,3).value="Teacher"





workbook.save(file)

