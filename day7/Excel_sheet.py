import openpyxl
file="C:\Automation Testing\BookData.xlsx"
workbook=openpyxl.load_workbook(file)
sheet=workbook["Sheet1"]

rows=sheet.max_row
colms=sheet.max_column
for r in range(1,rows+1) :
    for c in range(1,colms+1):
        print(sheet.cell(r,c).value,end='     ')
    print()
